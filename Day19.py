#reading a excel file and acccessing the values
import openpyxl
wb_obj = openpyxl.load_workbook("students.xlsx")
sheet_obj = wb_obj.active
cell_obj = sheet_obj.cell(row = 3, column = 2)
print(cell_obj.value)

#to read a specific student all record

import openpyxl
wb_obj = openpyxl.load_workbook("students.xlsx")
sheet_obj = wb_obj.active
for i in range(1,7):
    cell_obj = sheet_obj.cell(row=3, column=i)
    print(cell_obj.value)

#create a student management database

import mysql.connector
mydb=mysql.connector.connect(
    host="localhost",
    user="root",
    password="Aerub@1306"
)

dbse=mydb.cursor()
dbse.execute("CREATE DATABASE Student_Management")
dbse=mydb.cursor()
for entry in dbse:
    print(entry)

#created a student table in student_management database

import mysql.connector
mydb=mysql.connector.connect(
    host="localhost",
    user="root",
    password="Aerub@1306",
    database="student_management"
)
dbse=mydb.cursor()
dbse.execute("CREATE TABLE student (roll_no INT(10),Reg_no INT(10),NAME VARCHAR(255), semester1 INT(25),semester2 INT(25),semester3 INT(25), CGPA INT(35) ,PHONE_NUMBER INT,email_id VARCHAR(255))" )
dbse = mydb.cursor()

dbse.execute("SHOW TABLES")

for value in dbse:
  print(value)

import pandas as pd

df = pd.read_excel('students.xlsx')
import xlrd
import MySQLdb

xl_sheet = xlrd.open_workbook("students.xlsx")
xl_sheet
sheet_name = xl_sheet.sheet_names()
sheet_name
mydb = mysql.connector.connect(
    host="localhost",
    user="root",
    password="Aerub@1306",
    database="Students_Management"
)

cur = mydb.cursor()
for s in range(0, 1):
    sheet = xl_sheet.sheet_by_index(s)
    sql = "INSERT INTO student(roll_no,Reg_no,NAME,semester1,semester2 ,semester3 , CGPA ,email_id) VALUES(%s,%s,%s,%s,%s,%s,%s,%s)"
    for r in range(1, sheet.nrows):
        roll_no = sheet.cell(r, 0).value
        Reg_no = sheet.cell(r, 1).value
        NAME = sheet.cell(r, 2).value
        semester1 = sheet.cell(r, 3).value
        semester2 = sheet.cell(r, 4).value
        semester3 = sheet.cell(r, 5).value
        CGPA = sheet.cell(r, 6).value
        email_id = sheet.cell(r, 7).value
        values = (roll_no, Reg_no, NAME, semester1, semester2, semester3, CGPA, email_id)

        cur.execute(sql, values)
mydb.commit()
mycursor = mydb.cursor()

mycursor.execute("SELECT * FROM student3")

myresult = mycursor.fetchall()

for x in myresult:
    print(x)
mycursor = mydb.cursor()

mycursor.execute("SELECT NAME FROM student3 WHERE CGPA >6")

myresult = mycursor.fetchall()

for x in myresult:
    print(x)
mydb.commit()

mydb.close()
